"""Read sheets from the Excel database with openpyxl."""
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


def _is_empty_row(values) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


@lru_cache(maxsize=32)
def read_sheet(xlsx: str, sheet_name: str):
    """Return (headers, rows) where rows is a list of {header: value} dicts."""
    wb = load_workbook(Path(xlsx), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(it)
        except StopIteration:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]
        rows = []
        for values in it:
            if _is_empty_row(values):
                continue
            row = {}
            for h, v in zip(headers, values):
                if h:
                    row[h] = v
            rows.append(row)
        return [h for h in headers if h], rows
    finally:
        wb.close()
